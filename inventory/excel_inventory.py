#!/usr/bin/env python3

import json
import argparse
import re
from openpyxl import load_workbook
from pathlib import Path
import sys

def verify_file(path):
    return path.endswith('inventory')

def get_inventory(excel_filename, hostname_col, group_by_col, sheet_name=0):
    # Get absolute path of the Excel file (same directory as script)
    script_dir = Path(__file__).resolve().parent
    excel_path = script_dir / excel_filename

    if not excel_path.exists():
        print(f"Error: Excel file '{excel_path}' not found.")
        sys.exit(1)

    # Load workbook and sheet
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]

    # Extract headers
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    # Validate required columns
    if hostname_col not in headers or group_by_col not in headers:
        print(f"Error: Missing required columns '{hostname_col}' or '{group_by_col}' in Excel file.")
        sys.exit(1)

    inventory = {"_meta": {"hostvars": {}}}
    groups = {}

    hostname_idx = headers.index(hostname_col)
    group_idx = headers.index(group_by_col)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):  # skip empty rows
            continue

        hostname = str(row[hostname_idx]).strip() if row[hostname_idx] else None
        group_name = str(row[group_idx]).strip().split('_')[-1] if row[group_idx] else "ungrouped" 

        if not hostname:
            continue
        if "acs" in group_name:
           group_name = "ise"
        if "controller" in group_name:
           group_name = "wlc"
           
        # Add host to group
        if group_name not in groups:
            groups[group_name] = {"hosts": []}
        groups[group_name]["hosts"].append(hostname)

        # Add host variables
        host_vars = {}
        for i, header in enumerate(headers):
            cell_value = str(row[i]).strip().lower()
            if header not in [hostname_col, group_by_col]:
                if header == "servicenow.company": ##skip unneeded colums
                    continue
                elif header == "sn.u_anm_support_level": ##skip unneeded colums
                    continue
                elif header == "name":
                    host_vars["ansible_host"] = str(row[i])
                elif header == "auto.anm_os":
                    if "ios_xe" in cell_value:
                        host_vars["ansible_network_os"] = "ios"
                        host_vars["ansible_connection"] = "network_cli"
                    elif "ios_xr" in cell_value:
                        host_vars["ansible_network_os"] = "iosxr"
                        host_vars["ansible_connection"] = "network_cli"
                    elif "ios" in cell_value:
                        host_vars["ansible_network_os"] = "ios"
                        host_vars["ansible_connection"] = "network_cli"
                        host_vars["ansible_command_timeout"] = 180
                        host_vars["ansible_connect_timeout"] = 60
                    elif "nxos" in cell_value:
                        host_vars["ansible_network_os"] = "nxos"
                        host_vars["ansible_connection"] = "network_cli"
                    elif "asa" in cell_value:
                        host_vars["ansible_network_os"] = "asa"
                        host_vars["ansible_connection"] = "network_cli"
                    elif "ftd" in cell_value or "fxos" in cell_value:
                        host_vars["ansible_network_os"] = "ftd"
                        host_vars["ansible_httpapi_port"] = "443"
                        host_vars["ansible_httpapi_use_ssl"] = "True"      
                        host_vars["ansible_httpapi_validate_certs"] = "False" 
                        host_vars["ansible_fmc_verify"] = "False"
                    elif "fmc" in cell_value:
                        host_vars["ansible_network_os"] = "fmcansible"
                        host_vars["ansible_httpapi_port"] = "443"
                        host_vars["ansible_httpapi_use_ssl"] = "True"       
                        host_vars["ansible_httpapi_validate_certs"] = "False"
                        host_vars["ansible_fmc_verify"] = "False"
                    elif "ise" in cell_value:
                        host_vars["ansible_network_os"] = "ise"
                        host_vars["ansible_ise_verify"] = "False"
                    elif "pan-os" in cell_value:
                        host_vars["ansible_network_os"] = "panos"
                        host_vars["ansible_connection"] = "network_cli"
                    elif "fortios" in cell_value:
                        host_vars["ansible_network_os"] = "fortios"
                        host_vars["ansible_connection"] = "httpapi"
                    elif "sonicos" in cell_value:
                        host_vars["ansible_network_os"] = "sonicos"
                        host_vars["ansible_connection"] = "network_cli" 
                    elif "aireos" in cell_value:
                        host_vars["ansible_network_os"] = "aireos"
                        host_vars["ansible_connection"] = "network_cli"
                    else: 
                        host_vars["ansible_network_os"] = cell_value
                    host_vars["platform_os"] = cell_value                  
                elif "model" in header:
                    if any(x in cell_value for x in ("2960",)):
                        if any(x in cell_value for x in ("2960x", "2960xr")):
                            host_vars["platform_series"] = "c2960x"
                        if any(x in cell_value for x in ("2960cx",)):
                            host_vars["platform_series"] = "c2960cx"
                        if any(x in cell_value for x in ("2960l",)):
                            host_vars["platform_series"] = "c2960x"
                        elif any(x in cell_value for x in ("2960c", "2960cg")):
                            host_vars["platform_series"] = "c2960c"
                        elif any(x in cell_value for x in ("2960s",)):
                            host_vars["platform_series"] = "c2960s"
                    elif any(x in cell_value for x in ("c1000",)):
                        host_vars["platform_series"] = "c1000"
                    elif any(x in cell_value for x in ("3750",)):
                        host_vars["platform_series"] = "c3750x"
                    elif any(x in cell_value for x in ("3560",)):
                        if any(x in cell_value for x in ("cx",)):
                            host_vars["platform_series"] = "c3560cx"
                        elif any(x in cell_value for x in ("x",)):
                            host_vars["platform_series"] = "c3560x"
                        else:
                            host_vars["platform_series"] = "c3560"
                    elif any(x in cell_value for x in ("9200",)):
                        host_vars["platform_series"] = "c9200l"
                    elif any(x in cell_value for x in ("9300", "9400", "9500", "9600", "9407", "9410")):
                        host_vars["platform_series"] = "c9000"
                    elif any(x in cell_value for x in ("3650", "3850")):
                        host_vars["platform_series"] = "cat3k"
                    elif any(x in cell_value for x in ("8200", "8300")):
                        host_vars["platform_series"] = "c8000"
                    elif any(x in cell_value for x in ("8500",)):
                        if any(x in cell_value for x in ("l",)):
                            host_vars["platform_series"] = "c8500l"
                        elif any(x in cell_value for x in ("8500",)):
                            host_vars["platform_series"] = "c8500"
                        host_vars["platform_series"] = "c8000"
                    elif any(x in cell_value for x in ("isr",)):
                        if any(x in cell_value for x in ("4221",)):
                            host_vars["platform_series"] = "isr4200"
                        elif any(x in cell_value for x in ("4321", "4331", "4351")):
                            host_vars["platform_series"] = "isr4300"
                        elif any(x in cell_value for x in ("4431", "4451")):
                            host_vars["platform_series"] = "isr4400"
                    elif any(x in cell_value for x in ("asr",)):
                        if any(x in cell_value for x in ("1001", "1002-hx")):
                            host_vars["platform_series"] = "asr1000"
                        elif any(x in cell_value for x in ("1002-x",)):
                            host_vars["platform_series"] = "asr1002"
                        elif any(x in cell_value for x in ("1004",)):
                            host_vars["platform_series"] = "asr1004"
                        elif any(x in cell_value for x in ("1006-x",)):
                            host_vars["platform_series"] = "asr1006x"
                        elif any(x in cell_value for x in ("1006",)):
                            host_vars["platform_series"] = "asr1006"
                        elif any(x in cell_value for x in ("1009",)):
                            host_vars["platform_series"] = "asr1009"
                    elif any(x in cell_value for x in ("9800",)):
                        if any(x in cell_value for x in ("40",)):
                            host_vars["platform_series"] = "wlc9800-40"
                        elif any(x in cell_value for x in ("80",)):
                            host_vars["platform_series"] = "wlc9800-80"
                        elif any(x in cell_value for x in ("cl",)):
                            host_vars["platform_series"] = "wlc9800-cl"
                        elif any(x in cell_value for x in ("l",)):
                            host_vars["platform_series"] = "wlc9800-l"
                        elif any(x in cell_value for x in ("h",)):
                            host_vars["platform_series"] = "wlc9800-h"
                        elif any(x in cell_value for x in ("m",)):
                            host_vars["platform_series"] = "wlc9800-m"
                    elif any(x in cell_value for x in ("linux", "fmc")):
                        host_vars["platform_series"] = "FMCv"
                    elif cell_value.startswith("ws"):
                        clean_up = cell_value.split('-')
                        if any('2960xr' in s for s in clean_up):
                            host_vars["platform_series"] = "c2960x"
                            host_vars["ansible_command_timeout"] = 180
                            host_vars["ansible_connect_timeout"] = 60
                        else:
                            host_vars["platform_series"] = clean_up[1]
                    elif any(x in cell_value for x in ("threat defense", "ftd", "fpr")):
                        if any(x in cell_value for x in ("1010", "1120", "1140", "1150")):
                            host_vars["platform_series"] = "fp1k"
                        elif re.search(r"21\d\d", cell_value ):
                            host_vars["platform_series"] = "fp2k"
                        elif re.search(r"41\d\d", cell_value ):
                            host_vars["platform_series"] = "fp4k"
                        elif re.search(r"93\d\d", cell_value ):
                            host_vars["platform_series"] = "fp9k"
                        elif re.search(r"12\d\d", cell_value ):
                            host_vars["platform_series"] = "td1200"
                        elif re.search(r"31\d\d", cell_value ):
                            host_vars["platform_series"] = "fp3k"
                        elif re.search(r"42\d\d", cell_value ):
                            host_vars["platform_series"] = "td4200"
                        else:
                            host_vars["platform_series"] = "FTD"
                            host_vars["model"] = "FTD"
                    elif "panorama" in cell_value:
                        host_vars["platform_series"] = "panorama"
                    elif any(x in cell_value for x in ("pa","palo")):
                        if re.search(r"14\d\d", cell_value ):
                            host_vars["platform_series"] = "pa1400"
                        elif re.search(r"34\d\d", cell_value ):
                            host_vars["platform_series"] = "pa3400"
                        elif re.search(r"54\d\d", cell_value ):
                            host_vars["platform_series"] = "pa5400"
                        elif re.search(r"34\d\d", cell_value ):
                            host_vars["platform_series"] = "pa3400"
                        elif re.search(r"4\d\d", cell_value):
                            host_vars["platform_series"] = "pa400"
                        elif any(x in cell_value for x in ("vm",)):
                            host_vars["platform_series"] = "pa-vm"
                        else:
                            host_vars["platform_series"] = cell_value
                    elif any(x in cell_value for x in ("fgt","fort")):
                        host_vars["platform_series"] = cell_value
                    elif any(x in cell_value for x in ("sonic","sonicwall")):
                        if re.search(r"46\d\b", cell_value):
                            host_vars["platform_series"] = "nsa4600"
                        host_vars["platform_series"] = cell_value
                    elif row[i] is None:
                        host_vars["platform_series"] = "unknown"
                    else:
                        host_vars["platform_series"] = cell_value
                    host_vars["pid"] = cell_value
                elif row[i] is None:
                    host_vars[header] = "unknown"
                else:
                    host_vars[header] = cell_value
        inventory["_meta"]["hostvars"][hostname] = host_vars

        if host_vars["platform_series"] not in groups:
            groups[(host_vars["platform_series"])] = {"hosts": []}
        groups[(host_vars["platform_series"])]["hosts"].append(hostname)

    # Combine everything
    inventory.update(groups)
    print(json.dumps(inventory, indent=4))


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ansible Dynamic Inventory from Excel (no pandas, local path)")

    parser.add_argument("--hostname-col", default="displayName", help="Column containing hostnames")
    parser.add_argument("--group-by-col", default="sn.sys_class_name", help="Column to group hosts by")
    parser.add_argument("--sheet-name", default=0, help="Excel sheet name or index (default: 0)")
    parser.add_argument("--list", action="store_true", help="Show inventory (Ansible mode)")
    parser.add_argument("--file-name", default="inventory.xlsx", help="Excel filename in script directory")
    args = parser.parse_args()

    if args.list:
        get_inventory(args.file_name, args.hostname_col, args.group_by_col, args.sheet_name)
    else:
        get_inventory(args.file_name, args.hostname_col, args.group_by_col, args.sheet_name)